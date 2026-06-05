"""XLSX workbook reader tool for MCP server."""

import json
import logging
from pathlib import Path

from mcp.server import FastMCP
from mcp.types import TextContent
from openpyxl import load_workbook

from src.config import settings

logger = logging.getLogger(__name__)


def _validate_filename(filename: str) -> str | None:
    """Validate filename for path traversal. Returns error message or None."""
    if not filename:
        return "Error: filename is required"
    if "/" in filename or "\\" in filename or ".." in filename:
        return "Error: filename must not contain path separators or '..'"
    if not filename.lower().endswith(".xlsx"):
        return "Error: filename must end with .xlsx"
    return None


def xlsx_read_handler(server: FastMCP) -> None:
    """Register the xlsx_read tool."""

    @server.tool()
    async def xlsx_read(
        filename: str,
        sheet: str | None = None,
        max_rows: int = 1000,
    ):
        """Read an .xlsx workbook. Returns sheets with name, dimensions, headers, rows.

        sheet: read one sheet by name; omit for all. max_rows: cap per sheet (default 1000).
        """
        try:
            # --- Validate filename ---------------------------------------------
            err = _validate_filename(filename)
            if err:
                return [TextContent(type="text", text=err)]

            # --- Resolve file path ---------------------------------------------
            output_dir = Path(settings.FILE_OUTPUT_DIR)
            file_path = output_dir / filename

            if not file_path.exists():
                return [TextContent(
                    type="text",
                    text=f"Error: file '{filename}' not found in {settings.FILE_OUTPUT_DIR}",
                )]

            # --- Load workbook -------------------------------------------------
            wb = load_workbook(file_path, data_only=False)

            # --- Determine which sheets to read --------------------------------
            if sheet is not None:
                if sheet not in wb.sheetnames:
                    available = ", ".join(wb.sheetnames)
                    return [TextContent(
                        type="text",
                        text=f"Error: sheet '{sheet}' not found. Available sheets: {available}",
                    )]
                sheet_names = [sheet]
            else:
                sheet_names = wb.sheetnames

            # --- Read each sheet -----------------------------------------------
            sheets_data = []
            for sheet_name in sheet_names:
                ws = wb[sheet_name]
                dimensions = ws.dimensions or "A1"

                # Collect all rows up to max_rows + 1 (row 1 = headers)
                headers = []
                rows = []

                # Read headers from row 1
                if ws.max_row >= 1:
                    for col_idx in range(1, ws.max_column + 1):
                        cell = ws.cell(row=1, column=col_idx)
                        headers.append(cell.value)

                # Read data rows from row 2 onward
                row_count = 0
                for row_idx in range(2, ws.max_row + 1):
                    if row_count >= max_rows:
                        break
                    row_data = []
                    for col_idx in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        row_data.append(cell.value)
                    rows.append(row_data)
                    row_count += 1

                sheets_data.append({
                    "name": sheet_name,
                    "dimensions": dimensions,
                    "headers": headers,
                    "rows": rows,
                })

            # --- Build response ------------------------------------------------
            result = {
                "status": "success",
                "filename": filename,
                "sheets": sheets_data,
            }

            return [TextContent(type="text", text=json.dumps(result, indent=2))]

        except PermissionError as e:
            logger.error("Permission error reading file %s: %s", filename, str(e))
            msg = f"Error: permission denied reading {filename}"
            return [TextContent(type="text", text=f"{msg} - {str(e)}")]
        except Exception as e:
            logger.error("Unexpected error reading file %s: %s", filename, str(e))
            return [TextContent(type="text", text=f"Error: failed to read '{filename}' - {str(e)}")]

    logger.info("Registered xlsx_read tool")
