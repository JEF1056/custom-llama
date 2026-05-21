"""XLSX workbook creation tool for MCP server."""

import base64
import logging
import re
from io import BytesIO
from pathlib import Path

from mcp.server import FastMCP
from mcp.types import EmbeddedResource, TextContent
from openpyxl import Workbook
from openpyxl.chart.bar_chart import BarChart
from openpyxl.chart.line_chart import LineChart
from openpyxl.chart.pie_chart import PieChart
from openpyxl.chart.reference import Reference
from openpyxl.chart.scatter_chart import ScatterChart
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.config import settings

logger = logging.getLogger(__name__)

# Safety limit to prevent OOM in container
MAX_ROWS_PER_SHEET = 50_000

# Valid Excel sheet name: max 31 chars, cannot contain : \ / ? * [ ]
_INVALID_SHEET_CHARS = re.compile(r"[:\\/?\*\[\]]")

# Valid Excel range pattern, e.g. "A1:D10", "A1", "$A$1:$D$10"
_RANGE_RE = re.compile(r"^\$?[A-Z]{1,3}\$?\d+(:\$?[A-Z]{1,3}\$?\d+)?$")

# Valid chart types
_VALID_CHART_TYPES = {"bar", "line", "pie", "scatter"}

# Valid alignment values
_VALID_ALIGNMENTS = {"left", "center", "right"}


def _validate_filename(filename: str) -> str | None:
    """Validate filename for path traversal. Returns error message or None."""
    if not filename:
        return "Error: filename is required"
    if "/" in filename or "\\" in filename or ".." in filename:
        return "Error: filename must not contain path separators or '..'"
    if not filename.lower().endswith(".xlsx"):
        return "Error: filename must end with .xlsx"
    return None


def _validate_sheet_name(name: str) -> str | None:
    """Validate an Excel sheet name. Returns error message or None."""
    if not name:
        return "Error: sheet name is required"
    if len(name) > 31:
        return f"Error: sheet name '{name}' exceeds 31 characters"
    if _INVALID_SHEET_CHARS.search(name):
        return f"Error: sheet name '{name}' contains invalid characters (: \\ / ? * [ ])"
    return None


def _validate_range(range_str: str) -> str | None:
    """Validate an Excel range string. Returns error message or None."""
    if not _RANGE_RE.match(range_str.upper()):
        return f"Error: invalid Excel range '{range_str}'"
    return None


def _apply_formatting(ws, rule: dict, first_sheet_name: str) -> str | None:
    """Apply a single formatting rule to a worksheet. Returns error message or None."""
    range_str = rule.get("range")

    if not range_str:
        return "Error: formatting rule requires 'range'"

    err = _validate_range(range_str)
    if err:
        return err

    # Build style components
    font_kwargs = {}
    if rule.get("bold") is not None:
        font_kwargs["bold"] = rule["bold"]
    if rule.get("italic") is not None:
        font_kwargs["italic"] = rule["italic"]
    if rule.get("font_color") is not None:
        font_kwargs["color"] = rule["font_color"]

    fill = None
    if rule.get("fill") is not None:
        fill = PatternFill(start_color=rule["fill"], end_color=rule["fill"], fill_type="solid")

    align_kwargs = {}
    if rule.get("align") is not None:
        align_val = rule["align"]
        if align_val not in _VALID_ALIGNMENTS:
            valid = ", ".join(sorted(_VALID_ALIGNMENTS))
            return f"Error: invalid align '{align_val}', must be {valid}"
        align_kwargs["horizontal"] = align_val

    number_format = rule.get("number_format")

    # Apply to each cell in the range
    for row in ws[range_str]:
        for cell in row:
            if font_kwargs:
                cell.font = Font(**font_kwargs)
            if fill:
                cell.fill = fill
            if align_kwargs:
                cell.alignment = Alignment(**align_kwargs)
            if number_format:
                cell.number_format = number_format

    return None


def _add_chart(ws, chart_def: dict) -> str | None:
    """Add a chart to a worksheet. Returns error message or None."""
    chart_type = chart_def.get("type", "bar")
    if chart_type not in _VALID_CHART_TYPES:
        valid = ", ".join(sorted(_VALID_CHART_TYPES))
        return f"Error: invalid chart type '{chart_type}', must be {valid}"

    data_range = chart_def.get("data_range")
    if not data_range:
        return "Error: chart requires 'data_range'"

    err = _validate_range(data_range)
    if err:
        return err

    position = chart_def.get("position", "A1")
    err = _validate_range(position)
    if err:
        return err

    title = chart_def.get("title", "")

    # Create chart object
    chart_classes = {
        "bar": BarChart,
        "line": LineChart,
        "pie": PieChart,
        "scatter": ScatterChart,
    }
    chart = chart_classes[chart_type]()
    chart.title = title

    if chart_def.get("x_axis_title"):
        chart.x_axis.title = chart_def["x_axis_title"]
    if chart_def.get("y_axis_title"):
        chart.y_axis.title = chart_def["y_axis_title"]

    # Parse range to create Reference
    # data_range like "A1:D10" — openpyxl Reference needs sheet-prefixed range
    range_upper = data_range.upper().replace("$", "")
    sheet_qualified_range = f"'{ws.title}'!{range_upper}"

    # Use openpyxl Reference for the data range
    data = Reference(ws, range_string=sheet_qualified_range)
    chart.add_data(data, titles_from_data=True)

    # Place chart at the specified cell
    ws.add_chart(chart, position.upper().replace("$", ""))

    return None


def _auto_adjust_columns(ws) -> None:
    """Auto-adjust column widths based on cell content."""
    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_len = 0
        for cell in col_cells:
            if cell.value is not None:
                val_str = str(cell.value)
                # For formulas, use the formula string length
                max_len = max(max_len, len(val_str))
        # Add padding, set minimum width
        ws.column_dimensions[get_column_letter(col_idx)].width = max(max_len + 2, 10)


def xlsx_create_handler(server: FastMCP) -> None:
    """Register the xlsx_create tool."""

    @server.tool()
    async def xlsx_create(
        filename: str,
        sheets: list[dict],
        charts: list[dict] | None = None,
        formatting: list[dict] | None = None,
    ):
        """Create an Excel (.xlsx) workbook with sheets, formatting, and charts.

        Creates a valid .xlsx file with multiple sheets, optional formatting
        rules, and optional charts. Formulas are supported by prefixing cell
        values with ``=``.

        **Preferred workflow: create small, then edit**

        Rather than building a complete workbook in one call, create a minimal
        version first (e.g. headers + a few rows), then use ``xlsx_edit`` to
        add data, fix values, apply formatting, or add charts. This iterative
        approach is more reliable and easier to debug.

        **Sheet definition**

        Each sheet in the ``sheets`` list is a dict with:

        - ``name`` (str): Sheet name (default: "Sheet1"). Max 31 chars, no
          ``: \\ / ? * [ ]`` characters.
        - ``headers`` (list[str]): Column headers written to row 1.
        - ``rows`` (list[list]): Data rows. Each row is a list of values
          matching the header count. Use ``None`` for empty cells.
          Prefix a string with ``=`` to store it as a formula.

        **Formatting rules** (optional)

        Each rule in ``formatting`` is a dict with:

        - ``sheet`` (str): Target sheet name (default: first sheet).
        - ``range`` (str): Excel range, e.g. ``"A1:D1"``.
        - ``bold`` (bool): Bold font.
        - ``italic`` (bool): Italic font.
        - ``fill`` (str): Hex fill color, e.g. ``"4472C4"``.
        - ``font_color`` (str): Hex font color.
        - ``align`` (str): ``"left"``, ``"center"``, or ``"right"``.
        - ``number_format`` (str): Excel number format, e.g. ``"#,##0.00"``.

        **Charts** (optional)

        Each chart in ``charts`` is a dict with:

        - ``sheet`` (str): Target sheet name.
        - ``type`` (str): ``"bar"``, ``"line"``, ``"pie"``, or ``"scatter"``.
        - ``title`` (str): Chart title.
        - ``data_range`` (str): Excel data range, e.g. ``"A1:D10"``.
        - ``position`` (str): Cell to place chart, e.g. ``"F1"``.
        - ``x_axis_title`` (str, optional): X-axis label.
        - ``y_axis_title`` (str, optional): Y-axis label.

        The created file is returned as an MCP EmbeddedResource (base64 blob)
        so the LLM can read it directly, plus a TextContent summary.

        Args:
            filename: Output filename, must end with .xlsx (e.g., "report.xlsx")
            sheets: List of sheet definitions with headers and rows
            charts: Optional list of chart definitions
            formatting: Optional list of formatting rules

        Returns:
            MCP EmbeddedResource with base64-encoded .xlsx file, plus a
            TextContent summary with file details.
        """
        try:
            # --- Validate filename ---------------------------------------------
            err = _validate_filename(filename)
            if err:
                return [TextContent(type="text", text=err)]

            # --- Validate sheets -----------------------------------------------
            if not sheets:
                return [TextContent(type="text", text="Error: at least one sheet is required")]

            sheet_names = set()
            for i, sheet_def in enumerate(sheets):
                name = sheet_def.get("name", f"Sheet{i + 1}")
                err = _validate_sheet_name(name)
                if err:
                    return [TextContent(type="text", text=err)]
                if name in sheet_names:
                    return [TextContent(type="text", text=f"Error: duplicate sheet name '{name}'")]
                sheet_names.add(name)

                headers = sheet_def.get("headers", [])
                rows = sheet_def.get("rows", [])

                if len(rows) > MAX_ROWS_PER_SHEET:
                    msg = f"Error: sheet '{name}' has {len(rows)} rows"
                    return [TextContent(
                        type="text",
                        text=f"{msg}, exceeds limit of {MAX_ROWS_PER_SHEET}",
                    )]

                # Validate row lengths match headers
                for j, row in enumerate(rows):
                    if len(row) != len(headers):
                        msg = f"Error: sheet '{name}' row {j + 1} has {len(row)} cols"
                        return [TextContent(type="text", text=f"{msg}, expected {len(headers)}")]

            # --- Validate formatting rules -------------------------------------
            if formatting:
                for i, rule in enumerate(formatting):
                    if not rule.get("range"):
                        return [TextContent(type="text",
                            text=f"Error: formatting rule {i + 1} requires 'range'")]
                    err = _validate_range(rule["range"])
                    if err:
                        return [TextContent(type="text", text=err)]
                    if "sheet" in rule:
                        err = _validate_sheet_name(rule["sheet"])
                        if err:
                            return [TextContent(type="text", text=err)]

            # --- Validate charts -----------------------------------------------
            if charts:
                for i, chart_def in enumerate(charts):
                    if not chart_def.get("data_range"):
                        return [TextContent(type="text",
                            text=f"Error: chart {i + 1} requires 'data_range'")]
                    err = _validate_range(chart_def["data_range"])
                    if err:
                        return [TextContent(type="text", text=err)]
                    if "position" in chart_def:
                        err = _validate_range(chart_def["position"])
                        if err:
                            return [TextContent(type="text", text=err)]
                    if "sheet" in chart_def:
                        err = _validate_sheet_name(chart_def["sheet"])
                        if err:
                            return [TextContent(type="text", text=err)]

            # --- Create workbook -----------------------------------------------
            wb = Workbook()

            first_sheet_name = sheets[0].get("name", "Sheet1")

            for i, sheet_def in enumerate(sheets):
                name = sheet_def.get("name", f"Sheet{i + 1}")
                headers = sheet_def.get("headers", [])
                rows = sheet_def.get("rows", [])

                # Create or rename sheet
                if i == 0:
                    ws = wb.active
                    ws.title = name
                else:
                    ws = wb.create_sheet(title=name)

                # Write headers to row 1
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col_idx, value=header)

                # Write data rows
                for row_idx, row in enumerate(rows, 2):
                    for col_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if isinstance(value, str) and value.startswith("="):
                            # Store as formula
                            cell.value = value
                        else:
                            cell.value = value

            # --- Apply formatting ----------------------------------------------
            if formatting:
                for rule in formatting:
                    sheet_name = rule.get("sheet", first_sheet_name)
                    if sheet_name not in wb.sheetnames:
                        msg = f"Error: formatting references sheet '{sheet_name}'"
                        return [TextContent(type="text", text=f"{msg} which does not exist")]
                    ws = wb[sheet_name]
                    err = _apply_formatting(ws, rule, first_sheet_name)
                    if err:
                        return [TextContent(type="text", text=err)]

            # --- Add charts ----------------------------------------------------
            if charts:
                for chart_def in charts:
                    sheet_name = chart_def.get("sheet", first_sheet_name)
                    if sheet_name not in wb.sheetnames:
                        msg = f"Error: chart references sheet '{sheet_name}'"
                        return [TextContent(type="text", text=f"{msg} which does not exist")]
                    ws = wb[sheet_name]
                    err = _add_chart(ws, chart_def)
                    if err:
                        return [TextContent(type="text", text=err)]

            # --- Auto-adjust column widths for all sheets ----------------------
            for ws in wb.worksheets:
                _auto_adjust_columns(ws)

            # --- Save to buffer ------------------------------------------------
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            xlsx_bytes = output.getvalue()

            # --- Write to disk -------------------------------------------------
            output_dir = Path(settings.FILE_OUTPUT_DIR)
            output_dir.mkdir(parents=True, exist_ok=True)
            file_path = output_dir / filename
            file_path.write_bytes(xlsx_bytes)

            # --- Build response ------------------------------------------------
            b64_content = base64.b64encode(xlsx_bytes).decode("utf-8")
            resource_uri = f"file://{file_path}"

            embedded = EmbeddedResource(
                type="resource",
                resource={
                    "uri": resource_uri,
                    "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "blob": b64_content,
                },
                annotations={"priority": 1},
            )

            # Build summary
            sheet_summaries = []
            for ws in wb.worksheets:
                row_count = ws.max_row - 1  # subtract header row
                ncols = len(ws[1])
                sheet_summaries.append(
                    f"  - {ws.title}: {ncols} columns, {row_count} data rows")

            chart_count = len(charts) if charts else 0
            format_count = len(formatting) if formatting else 0

            info_text = (
                f"Excel workbook created successfully.\n"
                f"  Resource URI: {resource_uri}\n"
                f"  Download URL: {settings.FILE_BASE_URL}/files/{filename}\n"
                f"  Size: {len(xlsx_bytes)} bytes\n"
                f"  Sheets: {len(wb.worksheets)}\n"
                + "\n".join(sheet_summaries)
                + f"\n  Charts: {chart_count}\n"
                f"  Formatting rules: {format_count}\n"
                f"\n"
                f"The file is embedded above as a base64 MCP resource. "
                f"You can also access it later via the resource URI or download URL."
            )

            return [embedded, TextContent(type="text", text=info_text)]

        except PermissionError as e:
            logger.error("Permission error creating file %s: %s", filename, str(e))
            msg = f"Error: permission denied writing to {settings.FILE_OUTPUT_DIR}"
            return [TextContent(type="text", text=f"{msg} - {str(e)}")]
        except OSError as e:
            logger.error("OS error creating file %s: %s", filename, str(e))
            return [TextContent(type="text", text=f"Error: failed to create file - {str(e)}")]
        except Exception as e:
            logger.error("Unexpected error creating file %s: %s", filename, str(e))
            return [TextContent(type="text", text=f"Error: {str(e)}")]

    logger.info("Registered xlsx_create tool")
