"""XLSX workbook edit tool for MCP server."""

import base64
import logging
from io import BytesIO
from pathlib import Path

from mcp.server import FastMCP
from mcp.types import EmbeddedResource, TextContent
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from src.config import settings

logger = logging.getLogger(__name__)

# Reuse validation helpers from xlsx_create
_INVALID_SHEET_CHARS = __import__("re").compile(r"[:\\/?\*\[\]]")
_RANGE_RE = __import__("re").compile(r"^\$?[A-Z]{1,3}\$?\d+(:\$?[A-Z]{1,3}\$?\d+)?$")
_VALID_ALIGNMENTS = {"left", "center", "right"}


def _validate_filename(filename: str) -> str | None:
    """Validate filename for path traversal. Returns error message or None."""
    if not filename:
        return "Error: filename is required"
    if "/" in filename or "\\" in filename or ".." in filename:
        return "Error: filename must not contain path separators or '..'"
    if not filename.lower().endswith(".xlsx"):
        return "Error: filename must end with .xlsx"
    return None


def _validate_sheet_name(name: str) -> str | None:
    """Validate an Excel sheet name. Returns error message or None."""
    if not name:
        return "Error: sheet name is required"
    if len(name) > 31:
        return f"Error: sheet name '{name}' exceeds 31 characters"
    if _INVALID_SHEET_CHARS.search(name):
        return f"Error: sheet name '{name}' contains invalid characters (: \\ / ? * [ ])"
    return None


def _validate_range(range_str: str) -> str | None:
    """Validate an Excel range string. Returns error message or None."""
    if not _RANGE_RE.match(range_str.upper()):
        return f"Error: invalid Excel range '{range_str}'"
    return None


def xlsx_edit_handler(server: FastMCP) -> None:
    """Register the xlsx_edit tool."""

    @server.tool()
    async def xlsx_edit(
        filename: str,
        operations: list[dict],
    ):
        """Edit an .xlsx file with targeted operations (applied in order). Returns EmbeddedResource.

        Operations (each is a dict with "type"):
        - "update_cell": sheet, cell (e.g. "A1"), value (prefix = for formula)
        - "append_rows": sheet, rows [[...]]
        - "insert_rows": sheet, at_row (1-indexed), rows [[...]]
        - "delete_rows": sheet, from_row (1-indexed), count
        - "add_sheet": name, headers [...], rows (optional)
        - "rename_sheet": old_name, new_name
        - "delete_sheet": name
        - "format_range": sheet, range, bold, italic, fill (hex), font_color (hex),
          align (left/center/right), number_format
        - "clear_range": sheet, range (clears values, keeps formatting)
        """
        try:
            # --- Validate filename ---------------------------------------------
            err = _validate_filename(filename)
            if err:
                return [TextContent(type="text", text=err)]

            # --- Resolve file path ---------------------------------------------
            output_dir = Path(settings.FILE_OUTPUT_DIR)
            file_path = output_dir / filename

            if not file_path.exists():
                return [TextContent(
                    type="text",
                    text=f"Error: file '{filename}' not found in {settings.FILE_OUTPUT_DIR}",
                )]

            # --- Validate operations -------------------------------------------
            if not operations:
                return [TextContent(type="text", text="Error: at least one operation is required")]

            valid_types = {
                "update_cell", "append_rows", "insert_rows", "delete_rows",
                "add_sheet", "rename_sheet", "delete_sheet",
                "format_range", "clear_range",
            }
            for i, op in enumerate(operations):
                op_type = op.get("type")
                if op_type not in valid_types:
                    valid = ", ".join(sorted(valid_types))
                    return [TextContent(type="text",
                        text=f"Error: operation {i + 1} has invalid type '{op_type}', must be {valid}")]

            # --- Load workbook -------------------------------------------------
            wb = load_workbook(file_path, data_only=False)

            # --- Apply operations ----------------------------------------------
            changes = []

            for i, op in enumerate(operations):
                op_type = op["type"]
                sheet_name = op.get("sheet", wb.sheetnames[0])

                if sheet_name not in wb.sheetnames and op_type not in ("add_sheet",):
                    available = ", ".join(wb.sheetnames)
                    return [TextContent(type="text",
                        text=f"Error: operation {i + 1} references sheet '{sheet_name}' not found. Available: {available}")]

                # --- update_cell ---
                if op_type == "update_cell":
                    cell_ref = op.get("cell")
                    if not cell_ref:
                        return [TextContent(type="text",
                            text=f"Error: operation {i + 1} (update_cell) requires 'cell'")]
                    ws = wb[sheet_name]
                    cell = ws[cell_ref.upper().replace("$", "")]
                    value = op.get("value")
                    if isinstance(value, str) and value.startswith("="):
                        cell.value = value
                    else:
                        cell.value = value
                    changes.append(f"Updated {sheet_name}!{cell.coordinate} = {value!r}")

                # --- append_rows ---
                elif op_type == "append_rows":
                    rows = op.get("rows")
                    if not rows:
                        return [TextContent(type="text",
                            text=f"Error: operation {i + 1} (append_rows) requires 'rows'")]
                    ws = wb[sheet_name]
                    expected_cols = ws.max_column
                    for j, row in enumerate(rows):
                        if len(row) != expected_cols:
                            return [TextContent(type="text",
                                text=f"Error: operation {i + 1} row {j + 1} has {len(row)} cols, expected {expected_cols}")]
                        ws.append(row)
                    changes.append(f"Appended {len(rows)} rows to {sheet_name}")

                # --- insert_rows ---
                elif op_type == "insert_rows":
                    rows = op.get("rows")
                    at_row = op.get("at_row")
                    if not rows or at_row is None:
                        return [TextContent(type="text",
                            text=f"Error: operation {i + 1} (insert_rows) requires 'rows' and 'at_row'")]
                    ws = wb[sheet_name]
                    ws.insert_rows(at_row, amount=len(rows))
                    expected_cols = ws.max_column
                    for j, row in enumerate(rows):
                        if len(row) != expected_cols:
                            return [TextContent(type="text",
                                text=f"Error: operation {i + 1} row {j + 1} has {len(row)} cols, expected {expected_cols}")]
                        for k, value in enumerate(row, 1):
                            cell = ws.cell(row=at_row + j, column=k)
                            if isinstance(value, str) and value.startswith("="):
                                cell.value = value
                            else:
                                cell.value = value
                    changes.append(f"Inserted {len(rows)} rows at row {at_row} in {sheet_name}")

                # --- delete_rows ---
                elif op_type == "delete_rows":
                    from_row = op.get("from_row")
                    count = op.get("count")
                    if from_row is None or count is None:
                        return [TextContent(type="text",
                            text=f"Error: operation {i + 1} (delete_rows) requires 'from_row' and 'count'")]
                    ws = wb[sheet_name]
                    ws.delete_rows(from_row, amount=count)
                    changes.append(f"Deleted {count} rows from row {from_row} in {sheet_name}")

                # --- add_sheet ---
                elif op_type == "add_sheet":
                    name = op.get("name")
                    if not name:
                        return [TextContent(type="text",
                            text=f"Error: operation {i + 1} (add_sheet) requires 'name'")]
                    err = _validate_sheet_name(name)
                    if err:
                        return [TextContent(type="text", text=err)]
                    if name in wb.sheetnames:
                        return [TextContent(type="text",
                            text=f"Error: operation {i + 1} sheet '{name}' already exists")]
                    headers = op.get("headers", [])
                    rows = op.get("rows", [])
                    ws = wb.create_sheet(title=name)
                    for col_idx, header in enumerate(headers, 1):
                        ws.cell(row=1, column=col_idx, value=header)
                    for row_idx, row in enumerate(rows, 2):
                        for col_idx, value in enumerate(row, 1):
                            cell = ws.cell(row=row_idx, column=col_idx)
                            if isinstance(value, str) and value.startswith("="):
                                cell.value = value
                            else:
                                cell.value = value
                    changes.append(f"Added sheet '{name}' with {len(headers)} columns, {len(rows)} rows")

                # --- rename_sheet ---
                elif op_type == "rename_sheet":
                    old_name = op.get("old_name")
                    new_name = op.get("new_name")
                    if not old_name or not new_name:
                        return [TextContent(type="text",
                            text=f"Error: operation {i + 1} (rename_sheet) requires 'old_name' and 'new_name'")]
                    err = _validate_sheet_name(new_name)
                    if err:
                        return [TextContent(type="text", text=err)]
                    if old_name not in wb.sheetnames:
                        return [TextContent(type="text",
                            text=f"Error: operation {i + 1} sheet '{old_name}' not found")]
                    if new_name in wb.sheetnames:
                        return [TextContent(type="text",
                            text=f"Error: operation {i + 1} sheet '{new_name}' already exists")]
                    wb[old_name].title = new_name
                    changes.append(f"Renamed sheet '{old_name}' to '{new_name}'")

                # --- delete_sheet ---
                elif op_type == "delete_sheet":
                    name = op.get("name")
                    if not name:
                        return [TextContent(type="text",
                            text=f"Error: operation {i + 1} (delete_sheet) requires 'name'")]
                    if name not in wb.sheetnames:
                        return [TextContent(type="text",
                            text=f"Error: operation {i + 1} sheet '{name}' not found")]
                    if len(wb.sheetnames) <= 1:
                        return [TextContent(type="text",
                            text=f"Error: operation {i + 1} cannot delete the last sheet")]
                    del wb[name]
                    changes.append(f"Deleted sheet '{name}'")

                # --- format_range ---
                elif op_type == "format_range":
                    range_str = op.get("range")
                    if not range_str:
                        return [TextContent(type="text",
                            text=f"Error: operation {i + 1} (format_range) requires 'range'")]
                    err = _validate_range(range_str)
                    if err:
                        return [TextContent(type="text", text=err)]
                    ws = wb[sheet_name]

                    font_kwargs = {}
                    if op.get("bold") is not None:
                        font_kwargs["bold"] = op["bold"]
                    if op.get("italic") is not None:
                        font_kwargs["italic"] = op["italic"]
                    if op.get("font_color") is not None:
                        font_kwargs["color"] = op["font_color"]

                    fill = None
                    if op.get("fill") is not None:
                        fill = PatternFill(
                            start_color=op["fill"], end_color=op["fill"], fill_type="solid")

                    align_kwargs = {}
                    if op.get("align") is not None:
                        align_val = op["align"]
                        if align_val not in _VALID_ALIGNMENTS:
                            valid = ", ".join(sorted(_VALID_ALIGNMENTS))
                            return [TextContent(type="text",
                                text=f"Error: invalid align '{align_val}', must be {valid}")]
                        align_kwargs["horizontal"] = align_val

                    number_format = op.get("number_format")

                    for row in ws[range_str]:
                        for cell in row:
                            if font_kwargs:
                                cell.font = Font(**font_kwargs)
                            if fill:
                                cell.fill = fill
                            if align_kwargs:
                                cell.alignment = Alignment(**align_kwargs)
                            if number_format:
                                cell.number_format = number_format

                    changes.append(f"Formatted range {sheet_name}!{range_str}")

                # --- clear_range ---
                elif op_type == "clear_range":
                    range_str = op.get("range")
                    if not range_str:
                        return [TextContent(type="text",
                            text=f"Error: operation {i + 1} (clear_range) requires 'range'")]
                    err = _validate_range(range_str)
                    if err:
                        return [TextContent(type="text", text=err)]
                    ws = wb[sheet_name]
                    for row in ws[range_str]:
                        for cell in row:
                            cell.value = None
                    changes.append(f"Cleared range {sheet_name}!{range_str}")

            # --- Save to buffer ------------------------------------------------
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            xlsx_bytes = output.getvalue()

            # --- Write to disk -------------------------------------------------
            file_path.write_bytes(xlsx_bytes)

            # --- Build response ------------------------------------------------
            b64_content = base64.b64encode(xlsx_bytes).decode("utf-8")
            resource_uri = f"file://{file_path}"

            embedded = EmbeddedResource(
                type="resource",
                resource={
                    "uri": resource_uri,
                    "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "blob": b64_content,
                },
                annotations={"priority": 1},
            )

            # Build summary
            sheet_summaries = []
            for ws in wb.worksheets:
                row_count = ws.max_row - 1
                ncols = ws.max_column
                sheet_summaries.append(
                    f"  - {ws.title}: {ncols} columns, {row_count} data rows")

            info_text = (
                f"Excel workbook edited successfully.\n"
                f"  Resource URI: {resource_uri}\n"
                f"  Size: {len(xlsx_bytes)} bytes\n"
                f"  Sheets: {len(wb.worksheets)}\n"
                + "\n".join(sheet_summaries)
                + f"\n  Operations applied: {len(operations)}\n"
                + "\n".join(f"  - {c}" for c in changes)
                + "\n\n"
                "The file is embedded above as a base64 MCP resource. "
                "You can also access it later via the resource URI."
            )

            return [embedded, TextContent(type="text", text=info_text)]

        except PermissionError as e:
            logger.error("Permission error editing file %s: %s", filename, str(e))
            msg = f"Error: permission denied writing to {settings.FILE_OUTPUT_DIR}"
            return [TextContent(type="text", text=f"{msg} - {str(e)}")]
        except OSError as e:
            logger.error("OS error editing file %s: %s", filename, str(e))
            return [TextContent(type="text", text=f"Error: failed to edit file - {str(e)}")]
        except Exception as e:
            logger.error("Unexpected error editing file %s: %s", filename, str(e))
            return [TextContent(type="text", text=f"Error: {str(e)}")]

    logger.info("Registered xlsx_edit tool")
