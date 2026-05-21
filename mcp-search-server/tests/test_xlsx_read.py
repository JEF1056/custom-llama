"""Tests for the xlsx_read tool."""

import asyncio
import json
from pathlib import Path

import pytest
from openpyxl import Workbook

from src.config import settings
from src.tools.xlsx_read import xlsx_read_handler


class FakeServer:
    def __init__(self):
        self.tools = {}

    def tool(self, *args, **kwargs):
        def decorator(fn):
            self.tools[fn.__name__] = fn
            return fn
        return decorator


def _get_xlsx_read_fn():
    server = FakeServer()
    xlsx_read_handler(server)
    return server.tools["xlsx_read"]


def _run(fn, **kwargs):
    return asyncio.run(fn(**kwargs))


def _create_test_workbook(filename: str, sheets_data: list[dict]) -> Path:
    """Helper to create a test .xlsx file in the output directory."""
    wb = Workbook()
    for i, sheet_def in enumerate(sheets_data):
        if i == 0:
            ws = wb.active
            ws.title = sheet_def["name"]
        else:
            ws = wb.create_sheet(title=sheet_def["name"])

        for row_idx, row in enumerate(sheet_def["rows"], 1):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

    output_dir = Path(settings.FILE_OUTPUT_DIR)
    output_dir.mkdir(parents=True, exist_ok=True)
    file_path = output_dir / filename
    wb.save(file_path)
    return file_path


def test_xlsx_read_tool_registered():
    """Test that xlsx_read tool is registered in the server."""
    from src.server import create_server, register_tools

    server = create_server()
    register_tools(server)
    tool_names = [t.name for t in server._tool_manager.list_tools()]
    assert "xlsx_read" in tool_names


def test_xlsx_read_basic():
    """Test basic reading of a simple workbook."""
    _create_test_workbook("read_basic.xlsx", [
        {
            "name": "Sales",
            "rows": [
                ["Date", "Product", "Qty", "Total"],
                ["2026-01-01", "Widget", 10, 100],
                ["2026-01-02", "Gadget", 5, 75],
            ],
        },
    ])

    fn = _get_xlsx_read_fn()
    result = _run(fn, filename="read_basic.xlsx")

    assert len(result) == 1
    data = json.loads(result[0].text)
    assert data["status"] == "success"
    assert data["filename"] == "read_basic.xlsx"
    assert len(data["sheets"]) == 1

    sheet = data["sheets"][0]
    assert sheet["name"] == "Sales"
    assert sheet["headers"] == ["Date", "Product", "Qty", "Total"]
    assert len(sheet["rows"]) == 2
    assert sheet["rows"][0] == ["2026-01-01", "Widget", 10, 100]
    assert sheet["rows"][1] == ["2026-01-02", "Gadget", 5, 75]

    Path(settings.FILE_OUTPUT_DIR, "read_basic.xlsx").unlink()


def test_xlsx_read_multiple_sheets():
    """Test reading a workbook with multiple sheets."""
    _create_test_workbook("read_multi.xlsx", [
        {"name": "Sheet1", "rows": [["A", "B"], [1, 2]]},
        {"name": "Sheet2", "rows": [["X", "Y", "Z"], ["a", "b", "c"]]},
    ])

    fn = _get_xlsx_read_fn()
    result = _run(fn, filename="read_multi.xlsx")

    data = json.loads(result[0].text)
    assert len(data["sheets"]) == 2
    assert data["sheets"][0]["name"] == "Sheet1"
    assert data["sheets"][1]["name"] == "Sheet2"

    Path(settings.FILE_OUTPUT_DIR, "read_multi.xlsx").unlink()


def test_xlsx_read_specific_sheet():
    """Test reading a specific sheet by name."""
    _create_test_workbook("read_sheet.xlsx", [
        {"name": "Alpha", "rows": [["A"], [1]]},
        {"name": "Beta", "rows": [["B"], [2]]},
    ])

    fn = _get_xlsx_read_fn()
    result = _run(fn, filename="read_sheet.xlsx", sheet="Beta")

    data = json.loads(result[0].text)
    assert len(data["sheets"]) == 1
    assert data["sheets"][0]["name"] == "Beta"
    assert data["sheets"][0]["headers"] == ["B"]

    Path(settings.FILE_OUTPUT_DIR, "read_sheet.xlsx").unlink()


def test_xlsx_read_formulas():
    """Test that formulas are preserved as strings."""
    _create_test_workbook("read_formulas.xlsx", [
        {
            "name": "Calc",
            "rows": [
                ["Value", "Formula"],
                [10, "=A2*2"],
                [20, "=A3+A2"],
            ],
        },
    ])

    fn = _get_xlsx_read_fn()
    result = _run(fn, filename="read_formulas.xlsx")

    data = json.loads(result[0].text)
    sheet = data["sheets"][0]
    assert sheet["rows"][0] == [10, "=A2*2"]
    assert sheet["rows"][1] == [20, "=A3+A2"]

    Path(settings.FILE_OUTPUT_DIR, "read_formulas.xlsx").unlink()


def test_xlsx_read_max_rows():
    """Test that max_rows limits the number of rows returned."""
    rows = [[f"Row{i}"] for i in range(1, 201)]  # 200 data rows + 1 header
    _create_test_workbook("read_limit.xlsx", [
        {"name": "Data", "rows": [["Header"]] + rows},
    ])

    fn = _get_xlsx_read_fn()
    result = _run(fn, filename="read_limit.xlsx", max_rows=50)

    data = json.loads(result[0].text)
    sheet = data["sheets"][0]
    assert len(sheet["rows"]) == 50

    Path(settings.FILE_OUTPUT_DIR, "read_limit.xlsx").unlink()


def test_xlsx_read_missing_file():
    """Test error handling for missing files."""
    fn = _get_xlsx_read_fn()
    result = _run(fn, filename="nonexistent.xlsx")

    text = result[0].text
    assert "Error" in text
    assert "not found" in text


def test_xlsx_read_path_traversal():
    """Test that path traversal is blocked."""
    fn = _get_xlsx_read_fn()

    for bad_name in ["../etc/passwd.xlsx", "foo\\bar.xlsx", "..\\secret.xlsx"]:
        result = _run(fn, filename=bad_name)
        text = result[0].text
        assert "Error" in text


def test_xlsx_read_wrong_extension():
    """Test that non-.xlsx filenames are rejected."""
    fn = _get_xlsx_read_fn()
    result = _run(fn, filename="report.csv")

    text = result[0].text
    assert "Error" in text
    assert ".xlsx" in text


def test_xlsx_read_empty_filename():
    """Test that empty filename is rejected."""
    fn = _get_xlsx_read_fn()
    result = _run(fn, filename="")

    text = result[0].text
    assert "Error" in text


def test_xlsx_read_nonexistent_sheet():
    """Test error when requesting a sheet that doesn't exist."""
    _create_test_workbook("read_bad_sheet.xlsx", [
        {"name": "Only", "rows": [["A"], [1]]},
    ])

    fn = _get_xlsx_read_fn()
    result = _run(fn, filename="read_bad_sheet.xlsx", sheet="Missing")

    text = result[0].text
    assert "Error" in text
    assert "Missing" in text
    assert "Available" in text

    Path(settings.FILE_OUTPUT_DIR, "read_bad_sheet.xlsx").unlink()


def test_xlsx_read_dimensions():
    """Test that dimensions are returned."""
    _create_test_workbook("read_dims.xlsx", [
        {"name": "Data", "rows": [["A", "B"], [1, 2], [3, 4]]},
    ])

    fn = _get_xlsx_read_fn()
    result = _run(fn, filename="read_dims.xlsx")

    data = json.loads(result[0].text)
    sheet = data["sheets"][0]
    assert "dimensions" in sheet
    assert sheet["dimensions"]  # non-empty

    Path(settings.FILE_OUTPUT_DIR, "read_dims.xlsx").unlink()


def test_xlsx_read_empty_sheet():
    """Test reading a sheet with no data."""
    _create_test_workbook("read_empty.xlsx", [
        {"name": "Empty", "rows": []},
    ])

    fn = _get_xlsx_read_fn()
    result = _run(fn, filename="read_empty.xlsx")

    data = json.loads(result[0].text)
    sheet = data["sheets"][0]
    # Empty sheet: first row (if any) with all None values is not treated as header
    # (since `any(v is not None ...)` is False), so it lands in rows or headers is empty
    assert sheet["headers"] == [] or sheet["headers"] == [None]
    # rows may contain a [None] row from openpyxl's default cell
    assert all(v is None for row in sheet["rows"] for v in row) or sheet["rows"] == []

    Path(settings.FILE_OUTPUT_DIR, "read_empty.xlsx").unlink()


def test_xlsx_read_headers_only():
    """Test reading a sheet with headers but no data rows."""
    _create_test_workbook("read_headers.xlsx", [
        {"name": "Headers", "rows": [["Col1", "Col2", "Col3"]]},
    ])

    fn = _get_xlsx_read_fn()
    result = _run(fn, filename="read_headers.xlsx")

    data = json.loads(result[0].text)
    sheet = data["sheets"][0]
    assert sheet["headers"] == ["Col1", "Col2", "Col3"]
    assert sheet["rows"] == []

    Path(settings.FILE_OUTPUT_DIR, "read_headers.xlsx").unlink()


def test_xlsx_read_mixed_types():
    """Test reading cells with mixed types (str, int, float, None)."""
    _create_test_workbook("read_types.xlsx", [
        {
            "name": "Types",
            "rows": [
                ["Col1", "Col2", "Col3", "Col4", "Col5"],  # header row
                ["str", 42, 3.14, None, True],              # data row
            ],
        },
    ])

    fn = _get_xlsx_read_fn()
    result = _run(fn, filename="read_types.xlsx")

    data = json.loads(result[0].text)
    sheet = data["sheets"][0]
    assert sheet["headers"] == ["Col1", "Col2", "Col3", "Col4", "Col5"]
    row = sheet["rows"][0]
    assert row[0] == "str"
    assert row[1] == 42
    assert row[2] == 3.14
    assert row[3] is None
    assert row[4] is True

    Path(settings.FILE_OUTPUT_DIR, "read_types.xlsx").unlink()
