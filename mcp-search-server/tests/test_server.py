"""Tests for the MCP Search Server."""

import json
import os
import tempfile
from pathlib import Path

import pytest


def test_import():
    """Test that the server can be imported."""
    from src.server import create_server

    server = create_server()
    assert server is not None


def test_config():
    """Test that the config can be imported."""
    from src.config import settings

    assert settings.MCP_SERVER_HOST == "0.0.0.0"
    assert settings.MCP_SERVER_PORT == 3100
    assert settings.SEARCH_ENGINE == "duckduckgo"
    assert settings.MAX_RESULTS == 10


def test_search_models():
    """Test the search models."""
    from src.search.models import SearchResult, SearchResponse

    result = SearchResult(
        title="Test",
        url="https://example.com",
        snippet="Test snippet",
        engine="duckduckgo",
    )
    assert result.title == "Test"
    assert result.url == "https://example.com"
    assert result.snippet == "Test snippet"

    response = SearchResponse(query="test", results=[result], total=1)
    assert response.query == "test"
    assert len(response.results) == 1
    assert response.total == 1


def test_http_request_tool_registered():
    """Test that http_request tool is registered."""
    from src.server import create_server, register_tools

    server = create_server()
    register_tools(server)
    tool_names = [t.name for t in server._tool_manager.list_tools()]
    assert "http_request" in tool_names


def test_http_request_basic():
    """Test http_request with a simple GET."""
    import asyncio
    from src.tools.http_request import http_request_handler

    class FakeServer:
        def __init__(self):
            self.tools = {}

        def tool(self, *args, **kwargs):
            def decorator(fn):
                self.tools[fn.__name__] = fn
                return fn
            return decorator

    server = FakeServer()
    http_request_handler(server)
    fn = server.tools["http_request"]

    result_str = asyncio.run(fn("GET", "https://httpbin.org/get"))
    result = json.loads(result_str)
    assert result["status"] == "success"
    assert result["status_code"] == 200
    assert "elapsed_seconds" in result


def test_http_request_auth_bearer():
    """Test http_request with bearer auth."""
    import asyncio
    from src.tools.http_request import http_request_handler

    class FakeServer:
        def __init__(self):
            self.tools = {}

        def tool(self, *args, **kwargs):
            def decorator(fn):
                self.tools[fn.__name__] = fn
                return fn
            return decorator

    server = FakeServer()
    http_request_handler(server)
    fn = server.tools["http_request"]

    result_str = asyncio.run(fn(
        "GET",
        "https://httpbin.org/headers",
        auth={"type": "bearer", "value": "test-token"},
    ))
    result = json.loads(result_str)
    assert result["status"] == "success"
    assert result["status_code"] == 200


def test_http_request_unsupported_method():
    """Test http_request rejects unsupported methods."""
    import asyncio
    from src.tools.http_request import http_request_handler

    class FakeServer:
        def __init__(self):
            self.tools = {}

        def tool(self, *args, **kwargs):
            def decorator(fn):
                self.tools[fn.__name__] = fn
                return fn
            return decorator

    server = FakeServer()
    http_request_handler(server)
    fn = server.tools["http_request"]

    result_str = asyncio.run(fn("TRACE", "https://example.com"))
    result = json.loads(result_str)
    assert result["status"] == "error"


def test_time_now_tool_registered():
    """Test that time_now tool is registered."""
    from src.server import create_server, register_tools

    server = create_server()
    register_tools(server)
    tool_names = [t.name for t in server._tool_manager.list_tools()]
    assert "time_now" in tool_names


def test_time_now_default():
    """Test time_now returns current time."""
    import asyncio
    from src.tools.time_now import time_now_handler

    class FakeServer:
        def __init__(self):
            self.tools = {}

        def tool(self, *args, **kwargs):
            def decorator(fn):
                self.tools[fn.__name__] = fn
                return fn
            return decorator

    server = FakeServer()
    time_now_handler(server)
    fn = server.tools["time_now"]

    result_str = asyncio.run(fn())
    result = json.loads(result_str)
    assert result["status"] == "success"
    assert "iso_format" in result
    assert "date" in result
    assert "time" in result
    assert "day_of_week" in result


def test_time_now_timezone():
    """Test time_now with specific timezone."""
    import asyncio
    from src.tools.time_now import time_now_handler

    class FakeServer:
        def __init__(self):
            self.tools = {}

        def tool(self, *args, **kwargs):
            def decorator(fn):
                self.tools[fn.__name__] = fn
                return fn
            return decorator

    server = FakeServer()
    time_now_handler(server)
    fn = server.tools["time_now"]

    result_str = asyncio.run(fn(timezone_name="Asia/Tokyo"))
    result = json.loads(result_str)
    assert result["status"] == "success"
    assert result["timezone"] == "Asia/Tokyo"


def test_time_now_conversion():
    """Test time_now timezone conversion."""
    import asyncio
    from src.tools.time_now import time_now_handler

    class FakeServer:
        def __init__(self):
            self.tools = {}

        def tool(self, *args, **kwargs):
            def decorator(fn):
                self.tools[fn.__name__] = fn
                return fn
            return decorator

    server = FakeServer()
    time_now_handler(server)
    fn = server.tools["time_now"]

    result_str = asyncio.run(fn(
        timezone_name="America/New_York",
        convert_from_timezone="Asia/Tokyo",
        convert_from_time="2026-01-15 14:30:00",
    ))
    result = json.loads(result_str)
    assert result["status"] == "success"
    assert result["operation"] == "conversion"
    assert "converted_time" in result


def test_calculator_tool_registered():
    """Test that calculator tool is registered."""
    from src.server import create_server, register_tools

    server = create_server()
    register_tools(server)
    tool_names = [t.name for t in server._tool_manager.list_tools()]
    assert "calculator" in tool_names


def test_calculator_basic():
    """Test calculator basic arithmetic."""
    import asyncio
    from src.tools.calculator import calculator_handler

    class FakeServer:
        def __init__(self):
            self.tools = {}

        def tool(self, *args, **kwargs):
            def decorator(fn):
                self.tools[fn.__name__] = fn
                return fn
            return decorator

    server = FakeServer()
    calculator_handler(server)
    fn = server.tools["calculator"]

    result_str = asyncio.run(fn("2 + 3 * 4"))
    result = json.loads(result_str)
    assert result["status"] == "success"
    assert result["result"] == "14"


def test_calculator_symbolic():
    """Test calculator symbolic math."""
    import asyncio
    from src.tools.calculator import calculator_handler

    class FakeServer:
        def __init__(self):
            self.tools = {}

        def tool(self, *args, **kwargs):
            def decorator(fn):
                self.tools[fn.__name__] = fn
                return fn
            return decorator

    server = FakeServer()
    calculator_handler(server)
    fn = server.tools["calculator"]

    result_str = asyncio.run(fn("solve(x**2 - 4, x)"))
    result = json.loads(result_str)
    assert result["status"] == "success"


def test_calculator_matrix():
    """Test calculator matrix operations."""
    import asyncio
    from src.tools.calculator import calculator_handler

    class FakeServer:
        def __init__(self):
            self.tools = {}

        def tool(self, *args, **kwargs):
            def decorator(fn):
                self.tools[fn.__name__] = fn
                return fn
            return decorator

    server = FakeServer()
    calculator_handler(server)
    fn = server.tools["calculator"]

    result_str = asyncio.run(fn("Matrix([[1,2],[3,4]]).det()"))
    result = json.loads(result_str)
    assert result["status"] == "success"
    assert result["result"] == "-2"


def test_file_ops_tool_registered():
    """Test that file operations tools are registered."""
    from src.server import create_server, register_tools

    server = create_server()
    register_tools(server)
    tool_names = [t.name for t in server._tool_manager.list_tools()]
    assert "file_read" in tool_names
    assert "file_list" in tool_names
    assert "file_delete" in tool_names


def test_file_ops_create_and_read():
    """Test file create, read, and delete cycle."""
    import asyncio
    from src.tools.file_ops import file_operations_handler
    from src.tools.filetool import create_file_handler
    from src.config import settings

    class FakeServer:
        def __init__(self):
            self.tools = {}

        def tool(self, *args, **kwargs):
            def decorator(fn):
                self.tools[fn.__name__] = fn
                return fn
            return decorator

    server = FakeServer()
    create_file_handler(server)
    file_operations_handler(server)

    # Create a file
    create_fn = server.tools["create_file"]
    read_fn = server.tools["file_read"]
    list_fn = server.tools["file_list"]
    delete_fn = server.tools["file_delete"]

    # Create
    result = asyncio.run(create_fn(
        filename="test.txt",
        content="Hello, world!",
    ))
    assert len(result) == 2  # embedded resource + text

    # Read
    result_str = asyncio.run(read_fn("test.txt"))
    result = json.loads(result_str)
    assert result["status"] == "success"
    assert result["content"] == "Hello, world!"

    # List
    result_str = asyncio.run(list_fn())
    result = json.loads(result_str)
    assert result["status"] == "success"
    filenames = [e["name"] for e in result["entries"]]
    assert "test.txt" in filenames

    # Delete
    result_str = asyncio.run(delete_fn("test.txt"))
    result = json.loads(result_str)
    assert result["status"] == "success"


def test_file_ops_path_traversal_blocked():
    """Test that path traversal is blocked in file ops."""
    import asyncio
    from src.tools.file_ops import file_operations_handler

    class FakeServer:
        def __init__(self):
            self.tools = {}

        def tool(self, *args, **kwargs):
            def decorator(fn):
                self.tools[fn.__name__] = fn
                return fn
            return decorator

    server = FakeServer()
    file_operations_handler(server)
    read_fn = server.tools["file_read"]

    result_str = asyncio.run(read_fn("../etc/passwd"))
    result = json.loads(result_str)
    assert result["status"] == "error"


def test_code_run_tool_registered():
    """Test that code_run tool is registered."""
    from src.server import create_server, register_tools

    server = create_server()
    register_tools(server)
    tool_names = [t.name for t in server._tool_manager.list_tools()]
    assert "code_run" in tool_names


def test_code_run_basic():
    """Test code_run with simple code."""
    import asyncio
    from src.tools.code_run import code_run_handler

    class FakeServer:
        def __init__(self):
            self.tools = {}

        def tool(self, *args, **kwargs):
            def decorator(fn):
                self.tools[fn.__name__] = fn
                return fn
            return decorator

    server = FakeServer()
    code_run_handler(server)
    fn = server.tools["code_run"]

    result_str = asyncio.run(fn('print("hello")'))
    result = json.loads(result_str)
    assert result["status"] == "success"
    assert "hello" in result["stdout"]


def test_code_run_blocked_import():
    """Test that dangerous imports are blocked."""
    import asyncio
    from src.tools.code_run import code_run_handler

    class FakeServer:
        def __init__(self):
            self.tools = {}

        def tool(self, *args, **kwargs):
            def decorator(fn):
                self.tools[fn.__name__] = fn
                return fn
            return decorator

    server = FakeServer()
    code_run_handler(server)
    fn = server.tools["code_run"]

    result_str = asyncio.run(fn('import os'))
    result = json.loads(result_str)
    # Should error because os is blocked
    assert result["status"] == "error" or "ImportError" in result.get("stderr", "")


def test_content_extractor():
    """Test content extractor basic functionality."""
    from src.extractor.content import ContentExtractor

    extractor = ContentExtractor()
    html = """<html><head><title>Test</title></head>
    <body><p>Hello world</p><script>alert('xss')</script></body></html>"""
    result = extractor.extract(html)
    assert result["title"] == "Test"
    assert "Hello world" in result["content"]
    assert "alert" not in result["content"]


def test_content_extractor_boilerplate_removal():
    """Test that boilerplate elements are removed from the soup."""
    from src.extractor.content import ContentExtractor

    extractor = ContentExtractor()
    html = """<html><head><title>Test</title></head>
    <body>
      <nav><a href="/">Nav link</a></nav>
      <aside>Sidebar content</aside>
      <footer>Footer text</footer>
      <main><p>Main content</p></main>
      <div class="ad"><span>Ad text</span></div>
      <div class="cookie-banner">Accept cookies</div>
      <div class="sidebar">Sidebar widget</div>
      <div class="footer-widget">Widget</div>
      <div class="advertisement">Ad content</div>
    </body></html>"""
    result = extractor.extract(html)
    content = result["content"]

    # Structural elements should be removed
    assert "Nav link" not in content
    assert "Sidebar content" not in content
    assert "Footer text" not in content

    # Ad/boilerplate class elements should be removed
    assert "Ad text" not in content
    assert "Accept cookies" not in content
    assert "Sidebar widget" not in content
    assert "Widget" not in content
    assert "Ad content" not in content

    # Main content should be preserved
    assert "Main content" in content


def test_content_extractor_truncation():
    """Test content extractor truncation with token budget."""
    from src.extractor.content import ContentExtractor
    from src.config import settings

    extractor = ContentExtractor()
    # Content must exceed both max_length (char trigger) and token budget (~4 chars/token)
    # 4000 tokens * 4 = ~16000 chars; use 20000 to exceed budget
    html = "<html><body><p>" + "x" * 20000 + "</p></body></html>"
    result = extractor.extract(html, max_length=100, truncate=True)
    assert "Summarized" in result["content"]
    # Verify output respects token budget
    from src.extractor.content import ContentExtractor as CE
    estimated_tokens = CE._estimate_tokens(result["content"])
    assert estimated_tokens <= settings.FETCH_TOKEN_BUDGET
    # Result should fit within the token budget (~4000 tokens ≈ ~16000 chars)
    from src.config import settings
    assert len(result["content"]) <= settings.FETCH_TOKEN_BUDGET * 4


def test_summarize_token_budget_within_budget_returns_full():
    """Phase 1: content within token budget returns full content with code blocks."""
    from src.extractor.content import ContentExtractor

    extractor = ContentExtractor()
    # Small content that fits within budget
    html = "<html><body><p>Short paragraph.</p><pre><code>```python\nprint('hi')\n```</code></pre></body></html>"
    extractor._soup = None  # reset
    result = extractor.extract(html, max_length=100, truncate=True)
    # Content is small, should not be summarized
    assert "Summarized" not in result["content"]


def test_summarize_token_budget_exceeding_stays_within():
    """Phase 2/3: content exceeding token budget stays within budget."""
    from src.extractor.content import ContentExtractor

    extractor = ContentExtractor()
    # Build large content with headings
    paragraphs = "\n\n".join(f"Paragraph {i}: " + "word " * 50 for i in range(30))
    content = f"# Introduction\n\n{paragraphs}\n\n## Details\n\n" + "\n\n".join(
        f"Detail paragraph {i}: " + "text " * 40 for i in range(20)
    )
    headings = [
        {"level": 1, "text": "Introduction", "id": ""},
        {"level": 2, "text": "Details", "id": ""},
    ]
    # Simulate code blocks being extracted
    extractor._code_blocks = [
        {"language": "python", "content": "def hello():\n    print('world')\n" * 10}
    ]

    token_budget = 500
    summarized = extractor._summarize(
        content, headings, max_length=16000, original_length=len(content),
        token_budget=token_budget,
    )

    assert "Summarized" in summarized
    assert f"budget {token_budget} tokens" in summarized
    assert extractor._estimate_tokens(summarized) <= token_budget


def test_summarize_token_budget_preserves_code_blocks():
    """Summarized output preserves code blocks when exceeding budget."""
    from src.extractor.content import ContentExtractor

    extractor = ContentExtractor()
    paragraphs = "\n\n".join(f"Section {i} content: " + "data " * 60 for i in range(20))
    content = f"# Main\n\n{paragraphs}"
    headings = [{"level": 1, "text": "Main", "id": ""}]
    extractor._code_blocks = [
        {"language": "javascript", "content": "const x = 42;\nconsole.log(x);"},
        {"language": "python", "content": "import os\nprint(os.getcwd())"},
    ]

    token_budget = 400
    summarized = extractor._summarize(
        content, headings, max_length=16000, original_length=len(content),
        token_budget=token_budget,
    )

    # Code blocks should be preserved
    assert "```javascript" in summarized
    assert "```python" in summarized
    assert "const x = 42" in summarized
    assert "import os" in summarized
    # Should stay within budget
    assert extractor._estimate_tokens(summarized) <= token_budget


def test_summarize_char_fallback_when_no_token_budget():
    """When token_budget is None, falls back to char-based summarization."""
    from src.extractor.content import ContentExtractor

    extractor = ContentExtractor()
    content = "Some content " * 500
    headings = []

    result = extractor._summarize(
        content, headings, max_length=100, original_length=len(content),
        token_budget=None,
    )

    assert "Summarized" in result
    assert "chars" in result  # char-based header format
    assert len(result) <= 100


def test_summarize_token_budget_header_format():
    """Header uses token-based format when token_budget is provided."""
    from src.extractor.content import ContentExtractor

    extractor = ContentExtractor()
    content = "Word " * 2000
    headings = []
    extractor._code_blocks = []

    token_budget = 200
    result = extractor._summarize(
        content, headings, max_length=16000, original_length=len(content),
        token_budget=token_budget,
    )

    assert "[Summarized — original was" in result
    assert "tokens, budget" in result
    assert "chars" not in result  # should NOT use char-based format


def test_extract_main_content_targeting():
    """Verifies nav/aside/footer text is excluded when main content is targeted."""
    from src.extractor.content import ContentExtractor

    extractor = ContentExtractor()
    html = """<html><head><title>Page</title></head>
    <body>
      <nav><a href="/home">Home</a><a href="/about">About</a></nav>
      <aside>Related articles sidebar</aside>
      <main>
        <h1>Article Title</h1>
        <p>This is the actual article content.</p>
        <p>It has multiple paragraphs of real text.</p>
      </main>
      <footer>Copyright 2026. All rights reserved.</footer>
    </body></html>"""
    result = extractor.extract(html)
    content = result["content"]

    # Main content should be present
    assert "Article Title" in content
    assert "actual article content" in content
    assert "multiple paragraphs" in content

    # Structural boilerplate should be excluded
    assert "Home" not in content
    assert "About" not in content
    assert "Related articles sidebar" not in content
    assert "Copyright 2026" not in content


def test_extract_preserves_code_blocks():
    """Verifies code blocks from <pre><code> appear in output."""
    from src.extractor.content import ContentExtractor

    extractor = ContentExtractor()
    html = """<html><head><title>Code Page</title></head>
    <body><main>
      <p>Here is some code:</p>
      <pre><code>```python
def hello():
    print("Hello, world!")
```</code></pre>
      <p>And another:</p>
      <pre><code>```javascript
const x = 42;
console.log(x);
```</code></pre>
    </main></body></html>"""
    result = extractor.extract(html)

    # Code blocks should be extracted into _code_blocks
    assert len(extractor._code_blocks) == 2
    assert extractor._code_blocks[0]["language"] == "python"
    assert "hello" in extractor._code_blocks[0]["content"]
    assert extractor._code_blocks[1]["language"] == "javascript"
    assert "console.log" in extractor._code_blocks[1]["content"]

    # Reconstructed content should include the code blocks
    full = extractor._reconstruct_with_code_blocks(result["content"])
    assert "```python" in full
    assert "```javascript" in full


def test_token_estimation():
    """Verifies _estimate_tokens returns reasonable values."""
    from src.extractor.content import ContentExtractor
    import math

    extractor = ContentExtractor()

    # Empty string
    assert extractor._estimate_tokens("") == 0

    # "Hello world" = 11 chars → ceil(11/4) = 3
    assert extractor._estimate_tokens("Hello world") == 3

    # Exactly 4 chars → 1 token
    assert extractor._estimate_tokens("abcd") == 1

    # 5 chars → 2 tokens
    assert extractor._estimate_tokens("abcde") == 2

    # Large text: ~4 chars per token heuristic
    text = "x" * 16000
    assert extractor._estimate_tokens(text) == 4000

    # Verify formula matches implementation
    for length in [0, 1, 3, 4, 5, 100, 1000, 10000]:
        text = "a" * length
        expected = math.ceil(length / 4)
        assert extractor._estimate_tokens(text) == expected


def test_summarize_within_budget():
    """Verifies content within budget is returned in full (no summarization header)."""
    from src.extractor.content import ContentExtractor

    extractor = ContentExtractor()
    # Small content that easily fits within a generous budget
    content = "This is a short paragraph that should fit within the token budget without any summarization."
    headings = [{"level": 1, "text": "Short", "id": ""}]
    extractor._code_blocks = []

    token_budget = 1000  # generous budget
    result = extractor._summarize(
        content, headings, max_length=16000, original_length=len(content),
        token_budget=token_budget,
    )

    # Should NOT have summarization header — content returned in full
    assert "Summarized" not in result
    assert "short paragraph" in result


def test_summarize_exceeds_budget():
    """Verifies content exceeding budget is summarized and stays within budget."""
    from src.extractor.content import ContentExtractor

    extractor = ContentExtractor()
    # Build content that clearly exceeds a small budget
    paragraphs = "\n\n".join(
        f"Paragraph {i}: " + "word " * 100 for i in range(20)
    )
    content = paragraphs
    headings = []
    extractor._code_blocks = []

    token_budget = 200  # small budget that content will exceed
    result = extractor._summarize(
        content, headings, max_length=16000, original_length=len(content),
        token_budget=token_budget,
    )

    # Should have summarization header
    assert "Summarized" in result
    assert "budget" in result
    # Must stay within budget
    assert extractor._estimate_tokens(result) <= token_budget


def test_summarize_preserves_code_blocks():
    """Verifies code blocks survive summarization when budget is exceeded."""
    from src.extractor.content import ContentExtractor

    extractor = ContentExtractor()
    # Large text content to trigger summarization
    content = "\n\n".join(f"Section {i}: " + "data " * 80 for i in range(15))
    headings = []
    extractor._code_blocks = [
        {"language": "python", "content": "def foo():\n    return 42"},
        {"language": "bash", "content": "echo 'hello world'"},
    ]

    token_budget = 300
    result = extractor._summarize(
        content, headings, max_length=16000, original_length=len(content),
        token_budget=token_budget,
    )

    # Code blocks must be present in output
    assert "```python" in result
    assert "```bash" in result
    assert "def foo()" in result
    assert "echo 'hello world'" in result
    # Must stay within budget
    assert extractor._estimate_tokens(result) <= token_budget


def test_code_block_truncation():
    """Verifies oversized code blocks are truncated with [truncated] suffix."""
    from src.extractor.content import ContentExtractor

    extractor = ContentExtractor()
    # Create a code block that exceeds max_chars
    long_code = "line_of_code()  # " * 500  # ~9500 chars
    text = f"```python\n{long_code}\n```"

    max_chars = 1000
    blocks = extractor._extract_code_blocks(text, max_chars=max_chars)

    assert len(blocks) == 1
    assert blocks[0]["language"] == "python"
    assert len(blocks[0]["content"]) <= max_chars + len("[truncated]")
    assert "[truncated]" in blocks[0]["content"]

    # Small code block should NOT be truncated
    short_text = "```js\nconst x = 1;\n```"
    short_blocks = extractor._extract_code_blocks(short_text, max_chars=1000)
    assert len(short_blocks) == 1
    assert "[truncated]" not in short_blocks[0]["content"]
    assert "const x = 1" in short_blocks[0]["content"]


def test_all_tools_registered():
    """Test that all expected tools are registered."""
    from src.server import create_server, register_tools

    server = create_server()
    register_tools(server)
    tool_names = [t.name for t in server._tool_manager.list_tools()]

    expected = {
        "search", "fetch", "deep_search",
        "browser_create_session", "browser_navigate", "browser_screenshot",
        "browser_get_interactables",
        "browser_click", "browser_fill", "browser_evaluate",
        "browser_get_text", "browser_get_content", "browser_monitor",
        "browser_close", "browser_list_sessions",
        "create_file", "xlsx_create", "file_read", "file_list", "file_delete",
        "http_request", "code_run", "time_now", "calculator",
    }
    for name in expected:
        assert name in tool_names, f"Tool {name} not registered"


def test_xlsx_create_tool_registered():
    """Test that xlsx_create tool is registered."""
    from src.server import create_server, register_tools

    server = create_server()
    register_tools(server)
    tool_names = [t.name for t in server._tool_manager.list_tools()]
    assert "xlsx_create" in tool_names


def test_xlsx_create_basic():
    """Test xlsx_create creates a valid workbook with one sheet."""
    import asyncio
    import base64
    from io import BytesIO
    from openpyxl import load_workbook
    from mcp.types import EmbeddedResource
    from src.tools.xlsx_create import xlsx_create_handler

    class FakeServer:
        def __init__(self):
            self.tools = {}

        def tool(self, *args, **kwargs):
            def decorator(fn):
                self.tools[fn.__name__] = fn
                return fn
            return decorator

    server = FakeServer()
    xlsx_create_handler(server)
    fn = server.tools["xlsx_create"]

    result = asyncio.run(fn(
        filename="test_basic.xlsx",
        sheets=[{
            "name": "Data",
            "headers": ["Name", "Age", "City"],
            "rows": [
                ["Alice", 30, "New York"],
                ["Bob", 25, "London"],
            ],
        }],
    ))

    assert len(result) == 2
    embedded = result[0]
    assert isinstance(embedded, EmbeddedResource)
    assert embedded.resource.mimeType == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    # Verify the file is a valid .xlsx
    b64 = embedded.resource.blob
    wb = load_workbook(BytesIO(base64.b64decode(b64)))
    assert "Data" in wb.sheetnames
    ws = wb["Data"]
    assert ws.cell(row=1, column=1).value == "Name"
    assert ws.cell(row=2, column=1).value == "Alice"
    assert ws.cell(row=2, column=2).value == 30
    assert ws.cell(row=3, column=3).value == "London"


def test_xlsx_create_formulas():
    """Test xlsx_create stores formulas correctly."""
    import asyncio
    import base64
    from io import BytesIO
    from openpyxl import load_workbook
    from src.tools.xlsx_create import xlsx_create_handler

    class FakeServer:
        def __init__(self):
            self.tools = {}

        def tool(self, *args, **kwargs):
            def decorator(fn):
                self.tools[fn.__name__] = fn
                return fn
            return decorator

    server = FakeServer()
    xlsx_create_handler(server)
    fn = server.tools["xlsx_create"]

    result = asyncio.run(fn(
        filename="test_formulas.xlsx",
        sheets=[{
            "name": "Calc",
            "headers": ["A", "B", "Sum"],
            "rows": [
                [10, 20, "=A2+B2"],
                [5, 15, "=A3+B3"],
            ],
        }],
    ))

    embedded = result[0]
    wb = load_workbook(BytesIO(base64.b64decode(embedded.resource.blob)))
    ws = wb["Calc"]
    # Formulas should be stored as-is (starting with =)
    assert ws.cell(row=2, column=3).value == "=A2+B2"
    assert ws.cell(row=3, column=3).value == "=A3+B3"


def test_xlsx_create_multiple_sheets():
    """Test xlsx_create with multiple sheets."""
    import asyncio
    import base64
    from io import BytesIO
    from openpyxl import load_workbook
    from src.tools.xlsx_create import xlsx_create_handler

    class FakeServer:
        def __init__(self):
            self.tools = {}

        def tool(self, *args, **kwargs):
            def decorator(fn):
                self.tools[fn.__name__] = fn
                return fn
            return decorator

    server = FakeServer()
    xlsx_create_handler(server)
    fn = server.tools["xlsx_create"]

    result = asyncio.run(fn(
        filename="test_multi.xlsx",
        sheets=[
            {
                "name": "Sheet1",
                "headers": ["X"],
                "rows": [[1], [2]],
            },
            {
                "name": "Sheet2",
                "headers": ["Y"],
                "rows": [[3], [4]],
            },
        ],
    ))

    embedded = result[0]
    wb = load_workbook(BytesIO(base64.b64decode(embedded.resource.blob)))
    assert "Sheet1" in wb.sheetnames
    assert "Sheet2" in wb.sheetnames
    assert wb["Sheet1"].cell(row=2, column=1).value == 1
    assert wb["Sheet1"].cell(row=3, column=1).value == 2
    assert wb["Sheet2"].cell(row=2, column=1).value == 3
    assert wb["Sheet2"].cell(row=3, column=1).value == 4


def test_xlsx_create_path_traversal_blocked():
    """Test that path traversal is blocked in xlsx_create."""
    import asyncio
    from src.tools.xlsx_create import xlsx_create_handler

    class FakeServer:
        def __init__(self):
            self.tools = {}

        def tool(self, *args, **kwargs):
            def decorator(fn):
                self.tools[fn.__name__] = fn
                return fn
            return decorator

    server = FakeServer()
    xlsx_create_handler(server)
    fn = server.tools["xlsx_create"]

    # Test with path separator
    result = asyncio.run(fn(filename="../etc/passwd.xlsx", sheets=[{"headers": ["A"], "rows": []}]))
    assert len(result) == 1
    assert "Error" in result[0].text

    # Test with backslash
    result = asyncio.run(fn(filename="foo\\bar.xlsx", sheets=[{"headers": ["A"], "rows": []}]))
    assert len(result) == 1
    assert "Error" in result[0].text

    # Test with ..
    result = asyncio.run(fn(filename="..\\secret.xlsx", sheets=[{"headers": ["A"], "rows": []}]))
    assert len(result) == 1
    assert "Error" in result[0].text


def test_xlsx_create_invalid_sheet_name():
    """Test that invalid sheet names are rejected."""
    import asyncio
    from src.tools.xlsx_create import xlsx_create_handler

    class FakeServer:
        def __init__(self):
            self.tools = {}

        def tool(self, *args, **kwargs):
            def decorator(fn):
                self.tools[fn.__name__] = fn
                return fn
            return decorator

    server = FakeServer()
    xlsx_create_handler(server)
    fn = server.tools["xlsx_create"]

    # Sheet name with invalid character
    result = asyncio.run(fn(
        filename="test.xlsx",
        sheets=[{"name": "Bad:Name", "headers": ["A"], "rows": []}],
    ))
    assert len(result) == 1
    assert "Error" in result[0].text

    # Sheet name too long
    result = asyncio.run(fn(
        filename="test.xlsx",
        sheets=[{"name": "A" * 32, "headers": ["A"], "rows": []}],
    ))
    assert len(result) == 1
    assert "Error" in result[0].text

    # Duplicate sheet names
    result = asyncio.run(fn(
        filename="test.xlsx",
        sheets=[
            {"name": "Dup", "headers": ["A"], "rows": []},
            {"name": "Dup", "headers": ["B"], "rows": []},
        ],
    ))
    assert len(result) == 1
    assert "Error" in result[0].text


def test_xlsx_create_chart():
    """Test xlsx_create with a bar chart."""
    import asyncio
    import base64
    from io import BytesIO
    from openpyxl import load_workbook
    from src.tools.xlsx_create import xlsx_create_handler

    class FakeServer:
        def __init__(self):
            self.tools = {}

        def tool(self, *args, **kwargs):
            def decorator(fn):
                self.tools[fn.__name__] = fn
                return fn
            return decorator

    server = FakeServer()
    xlsx_create_handler(server)
    fn = server.tools["xlsx_create"]

    result = asyncio.run(fn(
        filename="test_chart.xlsx",
        sheets=[{
            "name": "Data",
            "headers": ["Category", "Value"],
            "rows": [
                ["A", 10],
                ["B", 20],
                ["C", 15],
            ],
        }],
        charts=[{
            "type": "bar",
            "title": "Test Chart",
            "data_range": "A1:B4",
            "position": "D1",
        }],
    ))

    embedded = result[0]
    wb = load_workbook(BytesIO(base64.b64decode(embedded.resource.blob)))
    ws = wb["Data"]
    # Chart should be added to the worksheet
    assert len(ws._charts) == 1
    chart = ws._charts[0]
    # Title is a Title object — verify it has text content
    assert chart.title is not None
    assert chart.title.tx is not None


def test_xlsx_create_formatting():
    """Test xlsx_create with formatting rules."""
    import asyncio
    import base64
    from io import BytesIO
    from openpyxl import load_workbook
    from src.tools.xlsx_create import xlsx_create_handler

    class FakeServer:
        def __init__(self):
            self.tools = {}

        def tool(self, *args, **kwargs):
            def decorator(fn):
                self.tools[fn.__name__] = fn
                return fn
            return decorator

    server = FakeServer()
    xlsx_create_handler(server)
    fn = server.tools["xlsx_create"]

    result = asyncio.run(fn(
        filename="test_format.xlsx",
        sheets=[{
            "name": "Data",
            "headers": ["Name", "Value"],
            "rows": [
                ["Alice", 100],
                ["Bob", 200],
            ],
        }],
        formatting=[{
            "range": "A1:B1",
            "bold": True,
            "fill": "4472C4",
            "font_color": "FFFFFF",
        }],
    ))

    embedded = result[0]
    wb = load_workbook(BytesIO(base64.b64decode(embedded.resource.blob)))
    ws = wb["Data"]
    # Header cells should have formatting applied
    assert ws.cell(row=1, column=1).font.bold is True
    assert ws.cell(row=1, column=1).fill.start_color.rgb == "004472C4" or ws.cell(row=1, column=1).fill.start_color.rgb == "4472C4"
