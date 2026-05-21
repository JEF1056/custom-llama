"""Tests for the xlsx_create tool."""

import asyncio
import base64
import json
import os
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.config import settings
from src.tools.xlsx_create import xlsx_create_handler


class FakeServer:
    def __init__(self):
        self.tools = {}

    def tool(self, *args, **kwargs):
        def decorator(fn):
            self.tools[fn.__name__] = fn
            return fn
        return decorator


def _get_xlsx_fn():
    server = FakeServer()
    xlsx_create_handler(server)
    return server.tools["xlsx_create"]


def _run(fn, **kwargs):
    return asyncio.run(fn(**kwargs))


def test_xlsx_create_tool_registered():
    """Test that xlsx_create tool is registered."""
    from src.server import create_server, register_tools

    server = create_server()
    register_tools(server)
    tool_names = [t.name for t in server._tool_manager.list_tools()]
    assert "xlsx_create" in tool_names


def test_xlsx_create_basic():
    """Test basic workbook creation with one sheet."""
    fn = _get_xlsx_fn()
    result = _run(fn,
        filename="basic.xlsx",
        sheets=[{
            "name": "Data",
            "headers": ["Name", "Age", "City"],
            "rows": [
                ["Alice", 30, "New York"],
                ["Bob", 25, "London"],
            ],
        }],
    )

    assert len(result) == 2  # EmbeddedResource + TextContent
    text_content = result[1]
    assert text_content.type == "text"
    assert "Excel workbook created successfully" in text_content.text

    # Verify the file was written to disk
    file_path = Path(settings.FILE_OUTPUT_DIR) / "basic.xlsx"
    assert file_path.exists()

    # Verify the file is a valid xlsx
    wb = load_workbook(file_path)
    assert "Data" in wb.sheetnames
    ws = wb["Data"]
    assert ws.cell(row=1, column=1).value == "Name"
    assert ws.cell(row=2, column=1).value == "Alice"
    assert ws.cell(row=2, column=2).value == 30
    assert ws.cell(row=3, column=2).value == 25

    # Cleanup
    file_path.unlink()


def test_xlsx_create_multiple_sheets():
    """Test workbook with multiple sheets."""
    fn = _get_xlsx_fn()
    result = _run(fn,
        filename="multi.xlsx",
        sheets=[
            {
                "name": "Sheet1",
                "headers": ["A", "B"],
                "rows": [["1", "2"]],
            },
            {
                "name": "Sheet2",
                "headers": ["X", "Y", "Z"],
                "rows": [["a", "b", "c"]],
            },
        ],
    )

    file_path = Path(settings.FILE_OUTPUT_DIR) / "multi.xlsx"
    wb = load_workbook(file_path)
    assert "Sheet1" in wb.sheetnames
    assert "Sheet2" in wb.sheetnames
    assert wb["Sheet1"].cell(row=1, column=1).value == "A"
    assert wb["Sheet2"].cell(row=1, column=3).value == "Z"

    file_path.unlink()


def test_xlsx_create_formulas():
    """Test that formulas are stored correctly."""
    fn = _get_xlsx_fn()
    result = _run(fn,
        filename="formulas.xlsx",
        sheets=[{
            "name": "Calc",
            "headers": ["Value", "Formula"],
            "rows": [
                [10, "=A2*2"],
                [20, "=A3+A2"],
            ],
        }],
    )

    file_path = Path(settings.FILE_OUTPUT_DIR) / "formulas.xlsx"
    wb = load_workbook(file_path)
    ws = wb["Calc"]
    # Formula cells should have the formula string (openpyxl keeps the leading =)
    assert ws.cell(row=2, column=2).value == "=A2*2"
    assert ws.cell(row=3, column=2).value == "=A3+A2"

    file_path.unlink()


def test_xlsx_create_formatting():
    """Test formatting rules are applied."""
    fn = _get_xlsx_fn()
    result = _run(fn,
        filename="formatted.xlsx",
        sheets=[{
            "name": "Data",
            "headers": ["Name", "Value"],
            "rows": [
                ["Alice", 100],
                ["Bob", 200],
            ],
        }],
        formatting=[
            {
                "range": "A1:B1",
                "bold": True,
                "fill": "4472C4",
                "font_color": "FFFFFF",
            },
            {
                "range": "B2:B3",
                "number_format": "#,##0.00",
            },
        ],
    )

    file_path = Path(settings.FILE_OUTPUT_DIR) / "formatted.xlsx"
    wb = load_workbook(file_path)
    ws = wb["Data"]

    # Header should be bold
    assert ws.cell(row=1, column=1).font.bold is True
    # Header should have fill
    assert ws.cell(row=1, column=1).fill.start_color.rgb == "004472C4" or ws.cell(row=1, column=1).fill.start_color.rgb == "4472C4"

    file_path.unlink()


def test_xlsx_create_charts():
    """Test chart creation."""
    fn = _get_xlsx_fn()
    result = _run(fn,
        filename="chart.xlsx",
        sheets=[{
            "name": "Data",
            "headers": ["Category", "Value"],
            "rows": [
                ["A", 10],
                ["B", 20],
                ["C", 30],
            ],
        }],
        charts=[{
            "sheet": "Data",
            "type": "bar",
            "title": "Test Chart",
            "data_range": "A1:B4",
            "position": "D1",
        }],
    )

    file_path = Path(settings.FILE_OUTPUT_DIR) / "chart.xlsx"
    wb = load_workbook(file_path)
    ws = wb["Data"]
    # Charts are stored in the worksheet's _charts list
    assert len(ws._charts) == 1
    chart = ws._charts[0]
    # chart.title is a Title object; extract text from rich text run
    assert "Test Chart" in str(chart.title)

    file_path.unlink()


def test_xlsx_create_chart_types():
    """Test different chart types."""
    fn = _get_xlsx_fn()
    for chart_type in ["bar", "line", "pie", "scatter"]:
        result = _run(fn,
            filename=f"chart_{chart_type}.xlsx",
            sheets=[{
                "name": "Data",
                "headers": ["X", "Y"],
                "rows": [[1, 2], [3, 4], [5, 6]],
            }],
            charts=[{
                "sheet": "Data",
                "type": chart_type,
                "title": f"{chart_type.title()} Chart",
                "data_range": "A1:B4",
                "position": "D1",
            }],
        )

        file_path = Path(settings.FILE_OUTPUT_DIR) / f"chart_{chart_type}.xlsx"
        wb = load_workbook(file_path)
        ws = wb["Data"]
        assert len(ws._charts) == 1, f"Chart type {chart_type} not created"
        file_path.unlink()


def test_xlsx_create_alignment():
    """Test alignment formatting."""
    fn = _get_xlsx_fn()
    result = _run(fn,
        filename="align.xlsx",
        sheets=[{
            "name": "Data",
            "headers": ["Left", "Center", "Right"],
            "rows": [["a", "b", "c"]],
        }],
        formatting=[
            {"range": "A1:A2", "align": "left"},
            {"range": "B1:B2", "align": "center"},
            {"range": "C1:C2", "align": "right"},
        ],
    )

    file_path = Path(settings.FILE_OUTPUT_DIR) / "align.xlsx"
    wb = load_workbook(file_path)
    ws = wb["Data"]
    assert ws.cell(row=1, column=1).alignment.horizontal == "left"
    assert ws.cell(row=1, column=2).alignment.horizontal == "center"
    assert ws.cell(row=1, column=3).alignment.horizontal == "right"

    file_path.unlink()


def test_xlsx_create_validation_path_traversal():
    """Test that path traversal is blocked."""
    fn = _get_xlsx_fn()

    for bad_name in ["../etc/passwd.xlsx", "foo\\bar.xlsx", "..\\secret.xlsx"]:
        result = _run(fn,
            filename=bad_name,
            sheets=[{"headers": ["A"], "rows": []}],
        )
        text = result[0].text
        assert "Error" in text


def test_xlsx_create_validation_no_extension():
    """Test that non-.xlsx filenames are rejected."""
    fn = _get_xlsx_fn()
    result = _run(fn,
        filename="report.csv",
        sheets=[{"headers": ["A"], "rows": []}],
    )
    text = result[0].text
    assert "Error" in text
    assert ".xlsx" in text


def test_xlsx_create_validation_empty_sheets():
    """Test that empty sheets list is rejected."""
    fn = _get_xlsx_fn()
    result = _run(fn,
        filename="empty.xlsx",
        sheets=[],
    )
    text = result[0].text
    assert "Error" in text


def test_xlsx_create_validation_duplicate_sheet_names():
    """Test that duplicate sheet names are rejected."""
    fn = _get_xlsx_fn()
    result = _run(fn,
        filename="dup.xlsx",
        sheets=[
            {"name": "Same", "headers": ["A"], "rows": []},
            {"name": "Same", "headers": ["B"], "rows": []},
        ],
    )
    text = result[0].text
    assert "Error" in text
    assert "duplicate" in text.lower()


def test_xlsx_create_validation_invalid_sheet_name():
    """Test that invalid sheet names are rejected."""
    fn = _get_xlsx_fn()
    result = _run(fn,
        filename="bad.xlsx",
        sheets=[{"name": "Invalid:Name", "headers": ["A"], "rows": []}],
    )
    text = result[0].text
    assert "Error" in text


def test_xlsx_create_validation_row_length_mismatch():
    """Test that row length must match header count."""
    fn = _get_xlsx_fn()
    result = _run(fn,
        filename="mismatch.xlsx",
        sheets=[{
            "headers": ["A", "B"],
            "rows": [["only_one"]],
        }],
    )
    text = result[0].text
    assert "Error" in text


def test_xlsx_create_validation_invalid_chart_type():
    """Test that invalid chart types are rejected."""
    fn = _get_xlsx_fn()
    result = _run(fn,
        filename="bad_chart.xlsx",
        sheets=[{"headers": ["A"], "rows": []}],
        charts=[{"type": "radar", "data_range": "A1:A2", "position": "B1"}],
    )
    text = result[0].text
    assert "Error" in text


def test_xlsx_create_validation_invalid_range():
    """Test that invalid Excel ranges are rejected."""
    fn = _get_xlsx_fn()
    result = _run(fn,
        filename="bad_range.xlsx",
        sheets=[{"headers": ["A"], "rows": []}],
        formatting=[{"range": "INVALID", "bold": True}],
    )
    text = result[0].text
    assert "Error" in text


def test_xlsx_create_default_sheet_name():
    """Test that sheets without name get default name."""
    fn = _get_xlsx_fn()
    result = _run(fn,
        filename="default_name.xlsx",
        sheets=[{
            "headers": ["A"],
            "rows": [["1"]],
        }],
    )

    file_path = Path(settings.FILE_OUTPUT_DIR) / "default_name.xlsx"
    wb = load_workbook(file_path)
    assert "Sheet1" in wb.sheetnames

    file_path.unlink()


def test_xlsx_create_empty_rows():
    """Test sheet with headers but no data rows."""
    fn = _get_xlsx_fn()
    result = _run(fn,
        filename="empty_rows.xlsx",
        sheets=[{
            "name": "Headers",
            "headers": ["A", "B", "C"],
            "rows": [],
        }],
    )

    file_path = Path(settings.FILE_OUTPUT_DIR) / "empty_rows.xlsx"
    wb = load_workbook(file_path)
    ws = wb["Headers"]
    assert ws.cell(row=1, column=1).value == "A"
    assert ws.cell(row=1, column=3).value == "C"
    assert ws.max_row == 1  # only header row

    file_path.unlink()


def test_xlsx_create_embedded_resource():
    """Test that the response contains a valid EmbeddedResource with base64 blob."""
    fn = _get_xlsx_fn()
    result = _run(fn,
        filename="resource_test.xlsx",
        sheets=[{
            "headers": ["A"],
            "rows": [["1"]],
        }],
    )

    embedded = result[0]
    assert embedded.type == "resource"
    resource = embedded.resource
    # resource is a BlobResourceContents object, access attributes directly
    assert hasattr(resource, "blob")
    assert resource.mimeType == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    # Verify blob is valid base64 and decodes to a valid xlsx
    decoded = base64.b64decode(resource.blob)
    wb = load_workbook(BytesIO(decoded))
    assert len(wb.sheetnames) == 1

    file_path = Path(settings.FILE_OUTPUT_DIR) / "resource_test.xlsx"
    if file_path.exists():
        file_path.unlink()


def test_xlsx_create_chart_axis_titles():
    """Test chart with axis titles."""
    fn = _get_xlsx_fn()
    result = _run(fn,
        filename="axis_chart.xlsx",
        sheets=[{
            "name": "Data",
            "headers": ["Month", "Sales"],
            "rows": [
                ["Jan", 100],
                ["Feb", 200],
                ["Mar", 150],
            ],
        }],
        charts=[{
            "sheet": "Data",
            "type": "line",
            "title": "Monthly Sales",
            "data_range": "A1:B4",
            "position": "D1",
            "x_axis_title": "Month",
            "y_axis_title": "Sales ($)",
        }],
    )

    file_path = Path(settings.FILE_OUTPUT_DIR) / "axis_chart.xlsx"
    wb = load_workbook(file_path)
    ws = wb["Data"]
    chart = ws._charts[0]
    # Axis titles are Title objects; check text is present in string representation
    assert "Month" in str(chart.x_axis.title)
    assert "Sales ($)" in str(chart.y_axis.title)

    file_path.unlink()


def test_xlsx_create_formatting_on_specific_sheet():
    """Test formatting applied to a non-first sheet."""
    fn = _get_xlsx_fn()
    result = _run(fn,
        filename="sheet_fmt.xlsx",
        sheets=[
            {"name": "First", "headers": ["A"], "rows": [["1"]]},
            {"name": "Second", "headers": ["B"], "rows": [["2"]]},
        ],
        formatting=[
            {"sheet": "Second", "range": "A1:A1", "bold": True, "italic": True},
        ],
    )

    file_path = Path(settings.FILE_OUTPUT_DIR) / "sheet_fmt.xlsx"
    wb = load_workbook(file_path)
    ws = wb["Second"]
    assert ws.cell(row=1, column=1).font.bold is True
    assert ws.cell(row=1, column=1).font.italic is True

    file_path.unlink()


def test_xlsx_create_formatting_references_nonexistent_sheet():
    """Test error when formatting references a sheet that doesn't exist."""
    fn = _get_xlsx_fn()
    result = _run(fn,
        filename="bad_fmt.xlsx",
        sheets=[{"name": "Only", "headers": ["A"], "rows": []}],
        formatting=[{"sheet": "Missing", "range": "A1:A1", "bold": True}],
    )
    text = result[0].text
    assert "Error" in text
    assert "Missing" in text


def test_xlsx_create_chart_references_nonexistent_sheet():
    """Test error when chart references a sheet that doesn't exist."""
    fn = _get_xlsx_fn()
    result = _run(fn,
        filename="bad_chart_sheet.xlsx",
        sheets=[{"name": "Only", "headers": ["A"], "rows": []}],
        charts=[{"sheet": "Missing", "type": "bar", "data_range": "A1:A2", "position": "B1"}],
    )
    text = result[0].text
    assert "Error" in text
    assert "Missing" in text
